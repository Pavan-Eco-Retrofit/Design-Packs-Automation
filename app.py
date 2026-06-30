import streamlit as st
import os
import glob
import re
import math
import requests
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import range_boundaries, column_index_from_string
import time
import tempfile
import io
import base64
from pathlib import Path

# ================================================================
# BRAND COLOURS (from Design Specifics logo)
# ================================================================
# Sage green:    #9CB5A8
# Dark navy:     #1F3A5F   (primary brand)
# Lavender blue: #8B95C9   (secondary accent)

# ================================================================
# SMARTSHEET COLUMN NAMES (climate risk / property flags)
# ================================================================
CONSERVATION_AREA_COL = "Conservation Area"
LISTED_BUILDING_COL   = "Listed Building"
RADON_RISK_COL        = "Radon Risk"
FLOOD_RISK_COL        = "Flood Risk"
SUBS_2030_COL         = "Subsidence 2030"
SUBS_2070_COL         = "Subsidence 2070"
WDR_RISK_COL          = "Wind Driven Rain Risk"
RAINFALL_RISK_COL     = "Rainfall Risk"
WIND_SPEED_RISK_COL   = "Wind speed Risk"
WILDFIRE_RISK_COL     = "Wildfire Risk"

# ================================================================
# PAGE CONFIG
# ================================================================
st.set_page_config(
    page_title="Power Asset · Retrofit Design Exporter",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)
# ================================================================
# LOGIN CREDENTIALS  —  EDIT HERE IN THE FUTURE
# ================================================================
# Emails and passwords are kept SEPARATE so you can change either
# one without touching the other.
#
#  • To rotate EVERYONE's password at once: change DEFAULT_PASSWORD.
#  • To give ONE user a different password:  put it as that user's
#    value in USERS (instead of None).
#  • To add / remove a user:                 add / remove a line in USERS.
#
# Emails are matched case-insensitively. Passwords are matched exactly.
# ----------------------------------------------------------------

DEFAULT_PASSWORD = "DSL@2026"

USERS = {
    # email (lower-case)                          : password (None = use DEFAULT_PASSWORD)
    "pavan.banavasi@designspecifics.co.uk":      None,
    "prasanth.tathireddy@designspecifics.co.uk": None,
    "gopi@compliancespecifics.co.uk":            "CSL@2026",
    "Noorie@compliancespecifics.co.uk"         : "CSL@2026",
    "deepak@compliancespecifics.co.uk"         : "CSL@2026
    "prashanth@compliancespecifics.co.uk": "CSL@2026",
    "harigopi@compliancespecifics.co.uk" : "CSL@2026",
    "sudheer.sivarathri@designspecifics.co.uk":  None,
    "srinu.dantala@designspecifics.co.uk":       None,
}


def get_password_for(email):
    """Return the effective password for an email (per-user override or default)."""
    pw = USERS.get(email)
    return pw if pw else DEFAULT_PASSWORD

# ================================================================
# LOAD LOGO
# Expects: template/Design_Specifics_logo.png
# ================================================================
def _load_logo_b64():
    candidates = [
        Path(__file__).parent / "template" / "Design_Specifics_logo.png",
        Path("template/Design_Specifics_logo.png"),
    ]

    for p in candidates:
        try:
            if p.exists():
                return base64.b64encode(p.read_bytes()).decode()
        except Exception:
            continue

    return None


LOGO_B64 = _load_logo_b64()

# ================================================================
# BUILT-IN FILES / SECRETS
# ================================================================
# Recommended: keep the token in .streamlit/secrets.toml as:
#SMARTSHEET_API_TOKEN = SMARTSHEET_API_TOKEN
# If you do not want to use secrets, paste the token between the quotes below.
def _get_smartsheet_token():
    try:
        token = st.secrets.get("SMARTSHEET_API_TOKEN", "")
    except Exception:
        token = ""
    return str(token).strip()

SMARTSHEET_API_TOKEN = _get_smartsheet_token()
#SMARTSHEET_API_TOKEN = "SMARTSHEET_API_TOKEN"  # Optional direct hardcode, not recommended for GitHub

# Put your Excel template here: assets/DESIGN EXPORT TEMPLATE.xlsx
BUILTIN_TEMPLATE_PATH = Path(__file__).parent / "template" / "DESIGN EXPORT TEMPLATE.xlsx"

def load_builtin_template_bytes():
    if not BUILTIN_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Built-in Excel template not found: {BUILTIN_TEMPLATE_PATH}"
        )
    return BUILTIN_TEMPLATE_PATH.read_bytes()

# ================================================================
# CUSTOM CSS — clean, minimalist, brand colours
# ================================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

:root {
    /* Brand */
    --brand-navy:    #1F3A5F;
    --brand-navy-2:  #2D4F7C;
    --brand-sage:    #9CB5A8;
    --brand-sage-2:  #B8CDC2;
    --brand-lav:     #8B95C9;
    --brand-lav-2:   #A8B1D6;

    /* Surfaces — light, soft, eye-friendly */
    --bg:        #F5F6F1;       /* warm off-white */
    --surface:   #FFFFFF;
    --surface2:  #FAFBF8;
    --border:    #E4E6DE;
    --border2:   #D2D5CB;

    /* Text */
    --text:      #1F3A5F;       /* matches brand navy */
    --text2:     #4A5A72;
    --muted:     #7A8294;
    --muted2:    #A0A6B3;

    /* States */
    --success:   #5B8F73;
    --warn:      #C99A4A;
    --danger:    #B85A5A;
}

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, sans-serif !important;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

.stApp { background: var(--bg) !important; }

/* Sidebar */
section[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
section[data-testid="stSidebar"] > div { padding-top: 1.5rem !important; }

/* Headings */
h1, h2, h3, h4 {
    font-family: 'Inter', sans-serif !important;
    color: var(--brand-navy) !important;
    letter-spacing: -0.01em !important;
    font-weight: 700 !important;
}

/* Logo container */
.logo-wrap {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.4rem 1.8rem;
    margin-bottom: 1.2rem;
    display: flex;
    align-items: center;
    gap: 1.5rem;
}
.logo-wrap img {
    height: 64px;
    width: auto;
    object-fit: contain;
}
.logo-divider {
    width: 1px;
    height: 50px;
    background: var(--border);
}
.logo-title {
    font-size: 1.3rem;
    font-weight: 700;
    color: var(--brand-navy);
    line-height: 1.2;
    letter-spacing: -0.01em;
}
.logo-subtitle {
    font-size: 0.85rem;
    color: var(--muted);
    margin-top: 2px;
    font-weight: 400;
}

/* Cards */
.card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.4rem 1.5rem;
    margin-bottom: 1.1rem;
    box-shadow: 0 1px 3px rgba(31, 58, 95, 0.03);
}

.card-label {
    font-size: 0.78rem;
    font-weight: 600;
    color: var(--brand-navy);
    margin-bottom: 0.6rem;
}

.card-hint {
    font-size: 0.8rem;
    color: var(--muted);
    margin-bottom: 1rem;
    line-height: 1.5;
}

/* Badges */
.badge {
    display: inline-block;
    background: #EAF0EC;
    border: 1px solid #C5D5CC;
    color: var(--success);
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 0.75rem;
    font-weight: 500;
}
.badge-brand {
    background: #EBEEF6;
    border-color: #C9D0E5;
    color: var(--brand-navy);
}
.badge-lav {
    background: #EFF1FA;
    border-color: #CFD4ED;
    color: #6671B3;
}
.badge-warn {
    background: #FAF1E1;
    border-color: #E8D5A8;
    color: var(--warn);
}

/* Log box */
.log-box {
    background: #FAFBF8;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 0.9rem 1.1rem;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem;
    color: var(--text2);
    min-height: 160px;
    max-height: 380px;
    overflow-y: auto;
    line-height: 1.7;
}
.log-ok    { color: var(--success); }
.log-warn  { color: var(--warn); }
.log-err   { color: var(--danger); }
.log-info  { color: var(--brand-navy-2); }
.log-empty {
    color: var(--muted2);
    font-style: italic;
    font-family: 'Inter', sans-serif;
}

/* Primary button — brand navy */
.stButton > button {
    background: var(--brand-navy) !important;
    color: #FFFFFF !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    border: 1px solid var(--brand-navy) !important;
    border-radius: 8px !important;
    padding: 0.65rem 1.5rem !important;
    transition: all 0.15s ease !important;
    width: 100%;
    box-shadow: 0 1px 2px rgba(31, 58, 95, 0.1);
}
.stButton > button:hover:not(:disabled) {
    background: var(--brand-navy-2) !important;
    border-color: var(--brand-navy-2) !important;
    box-shadow: 0 4px 12px rgba(31, 58, 95, 0.18) !important;
    transform: translateY(-1px);
}
.stButton > button:disabled {
    background: #CFD3C9 !important;
    border-color: #CFD3C9 !important;
    color: #FFFFFF !important;
    cursor: not-allowed !important;
    box-shadow: none !important;
}

/* Download buttons */
[data-testid="stDownloadButton"] > button {
    background: var(--surface) !important;
    color: var(--brand-navy) !important;
    border: 1.5px solid var(--brand-sage) !important;
    font-weight: 600 !important;
    box-shadow: none !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #EAF0EC !important;
    border-color: var(--brand-navy) !important;
}

/* File uploader */
div[data-testid="stFileUploader"] {
    background: var(--surface) !important;
    border: 1.5px dashed var(--border2) !important;
    border-radius: 8px !important;
    padding: 0.4rem !important;
    transition: border-color 0.15s !important;
}
div[data-testid="stFileUploader"]:hover {
    border-color: var(--brand-sage) !important;
}
div[data-testid="stFileUploader"] section { background: transparent !important; }
div[data-testid="stFileUploader"] section > button {
    background: var(--surface2) !important;
    color: var(--brand-navy) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
    font-weight: 500 !important;
}
div[data-testid="stFileUploader"] section > button:hover {
    border-color: var(--brand-navy) !important;
}

/* Inputs */
.stTextInput > div > div > input,
.stSelectbox > div > div,
.stNumberInput > div > div > input,
.stMultiSelect > div > div {
    background: var(--surface) !important;
    border: 1px solid var(--border2) !important;
    color: var(--text) !important;
    border-radius: 7px !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.9rem !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-color: var(--brand-navy) !important;
    box-shadow: 0 0 0 3px rgba(31, 58, 95, 0.08) !important;
}

/* Labels */
.stTextInput label, .stSelectbox label, .stFileUploader label,
.stMultiSelect label, .stNumberInput label {
    color: var(--text2) !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    text-transform: none !important;
    margin-bottom: 0.3rem !important;
}

/* Help text & captions */
.stCaption, [data-testid="stCaptionContainer"] {
    color: var(--muted) !important;
    font-size: 0.8rem !important;
}

/* Horizontal rules */
hr {
    border: none !important;
    border-top: 1px solid var(--border) !important;
    margin: 1rem 0 !important;
}

/* Progress bar */
.stProgress > div > div > div { background: var(--brand-navy) !important; }

/* Checklist items */
.check-item {
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 8px 12px;
    margin-bottom: 6px;
    font-size: 0.85rem;
    color: var(--text2);
    display: flex;
    align-items: center;
    gap: 8px;
}
.check-item.ok { background: #F1F6F2; border-color: #D5E1D7; }

[data-testid="column"] { padding: 0 0.4rem !important; }

.stAlert {
    border-radius: 8px !important;
    border: 1px solid var(--border) !important;
}

/* Output mini-card */
.output-mini {
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 1rem;
    text-align: center;
    margin-bottom: 0.6rem;
}
.output-mini-icon { font-size: 1.6rem; }
.output-mini-title {
    font-weight: 600;
    color: var(--brand-navy);
    margin-top: 0.2rem;
}
.output-mini-sub {
    font-size: 0.75rem;
    color: var(--muted);
    margin-top: 0.25rem;
}

/* Login */
.login-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 2rem 2.2rem;
    margin-top: 1rem;
    box-shadow: 0 4px 18px rgba(31, 58, 95, 0.06);
}
.login-title {
    font-size: 1.15rem;
    font-weight: 700;
    color: var(--brand-navy);
    margin-bottom: 0.3rem;
}
.login-sub {
    font-size: 0.85rem;
    color: var(--muted);
    margin-bottom: 1.2rem;
}
</style>
""", unsafe_allow_html=True)

# ================================================================
# LOGIN GATE
# Renders a styled sign-in form. Until the user signs in, the rest
# of the app does not run (st.stop()). Each browser session is
# independent, so one person logging in / uploading never affects
# another person's session.
# ================================================================
def render_login():
    if LOGO_B64:
        st.markdown(
            f'''
            <div class="logo-wrap">
                <img src="data:image/png;base64,{LOGO_B64}" alt="Design Specifics" />
                <div class="logo-divider"></div>
                <div>
                    <div class="logo-title">Power Asset · Retrofit Design Exporter</div>
                    <div class="logo-subtitle">Please sign in to continue</div>
                </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '''
            <div class="logo-wrap">
                <div>
                    <div class="logo-title">Power Asset · Retrofit Design Exporter</div>
                    <div class="logo-subtitle">Please sign in to continue</div>
                </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )

    _l, mid, _r = st.columns([1, 1.3, 1])
    with mid:
        st.markdown('<div class="login-card">', unsafe_allow_html=True)
        st.markdown('<div class="login-title">Sign in</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-sub">Use your Design Specifics account.</div>', unsafe_allow_html=True)

        with st.form("login_form", clear_on_submit=False):
            email_in = st.text_input("Email", placeholder="name@designspecifics.co.uk")
            pass_in  = st.text_input("Password", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Sign in")

        if submitted:
            email_norm = (email_in or "").strip().lower()
            if email_norm in USERS and (pass_in or "") == get_password_for(email_norm):
                st.session_state["authenticated"] = True
                st.session_state["user_email"] = email_norm
                st.rerun()
            else:
                st.error("Invalid email or password.")

        st.markdown('</div>', unsafe_allow_html=True)


if not st.session_state.get("authenticated"):
    render_login()
    st.stop()


# ================================================================
# CUSTOM ROUND
# ================================================================
def custom_round(x) -> float:
    try:
        x = float(x)
    except (TypeError, ValueError):
        return x
    floor_val = math.floor(x)
    return float(math.ceil(x) if (x - floor_val) >= 0.5 else floor_val)

# ================================================================
# ROOM CLASSIFICATION & SORT
# ================================================================
WET_ROOM_KEYWORDS = ["kitchen", "bathroom", "wetroom", "wet room", "wc", "w/c", "toilet", "utility"]
WET_ROOM_PRIORITY = {
    "kitchen": 0, "utility": 1, "bathroom": 2,
    "wetroom": 3, "wet room": 3, "wc": 4, "w/c": 4, "toilet": 5,
}

def is_wet_room(room_name): return any(kw in room_name.lower() for kw in WET_ROOM_KEYWORDS)

def is_wc_or_utility(room_name):
    return any(kw in room_name.lower().strip() for kw in ["wc", "w/c", "toilet", "utility"])

def is_bathroom_or_kitchen(room_name):
    name = room_name.lower()
    return "kitchen" in name or "bathroom" in name

def wet_room_sort_key(room_name):
    name = room_name.lower()
    for kw, priority in WET_ROOM_PRIORITY.items():
        if kw in name:
            return (0, priority)
    if any(kw in name for kw in ["living", "lounge", "dining", "reception"]): return (1, 0)
    if "bedroom" in name: return (2, 0)
    return (3, 0)

# ================================================================
# PDF HELPERS
# ================================================================
def read_pdf_text(path):
    with fitz.open(path) as doc:
        return "\n".join(page.get_text() for page in doc)

def read_pdf_page_lines(path, page_num):
    with fitz.open(path) as doc:
        if page_num < 0 or page_num >= len(doc):
            return []
        return (doc[page_num].get_text() or "").splitlines()

# ================================================================
# ADDRESS EXTRACTION
# ================================================================
def extract_property_address_from_cr(path):
    lines = [l.strip() for l in read_pdf_page_lines(path, 0) if l.strip()]
    addr_lines = []
    for i, line in enumerate(lines):
        if "Property Address" in line:
            for l in lines[i + 1: i + 7]:
                if any(x in l.lower() for x in ["assessment", "reference", "date"]):
                    break
                addr_lines.append(l)
            break
    addr = " ".join(addr_lines)
    addr = re.sub(r"\b([A-Z]{1,2}\d{1,2}[A-Z]?)\s+(\d[A-Z]{2})\b", r"\1 \2", addr, flags=re.I)
    return re.sub(r"\s{2,}", " ", addr).strip()

def extract_postcode(text):
    m = re.search(r"[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}", str(text), re.I)
    return m.group(0).upper() if m else ""

def extract_house_number(address):
    if pd.isna(address): return None
    address = str(address).strip()
    m = re.search(r"(?i)(?:flat|apartment)\s+([A-Za-z0-9]{1,5})", address)
    if m: return m.group(1)
    m = re.match(r"^([A-Za-z]?\d+[A-Za-z]?)\b", address)
    if m: return m.group(1)
    m = re.search(r"\b(\d+[A-Za-z]?)\b", address)
    return m.group(1) if m else None

def extract_street_number(address, house_number):
    if not address or not house_number: return None
    remaining = re.sub(rf"\b{re.escape(str(house_number))}\b", "", address, count=1, flags=re.I)
    m = re.search(r"\b(\d+[A-Za-z]?)\b", remaining)
    return m.group(1) if m else None

def normalize_text(x):
    if pd.isna(x): return ""
    return re.sub(r"[^a-z0-9]", "", str(x).lower())

def normalize_postcode(x):
    if pd.isna(x): return ""
    return re.sub(r"\s+", "", str(x).lower())

# ================================================================
# TENURE NORMALISATION
# ================================================================
def normalise_tenure(raw_tenure):
    if not raw_tenure: return raw_tenure
    cleaned = raw_tenure.strip()
    if re.search(r"rent(?:al|ed)\s*\(?\s*social\s*\)?", cleaned, re.I):
        return "Social Housing"
    return cleaned

# ================================================================
# EXTRACT PROPERTY DETACHMENT FROM CR PDF PAGE 3
# ================================================================
def extract_property_detachment_from_cr(path):
    lines = read_pdf_page_lines(path, 2)
    for line in lines:
        stripped = line.strip()
        if stripped.lower().startswith("property detachment"):
            value = re.sub(r"(?i)property\s+detachment\s*", "", stripped).strip()
            if value: return value
    return ""

# ================================================================
# EXTRACT PROPERTY TYPE FROM EPR PDF
# Pulls the "Property Type" value from the EPR dwelling-details
# block (the line right above "Total Floor Area"),
# e.g. "End-Terrace House", "Mid-Terrace House", "Detached Bungalow".
# ================================================================
def extract_property_type_from_epr(epr_text):
    if not epr_text:
        return ""

    # 1) Same-line: "Property Type   End-Terrace House"
    #    Only accept when the captured value looks like a dwelling type
    #    (otherwise we may be matching a label-only line in a stacked layout).
    type_keywords = ("house", "flat", "maisonette", "bungalow", "park home")
    m = re.search(
        r"Property\s*Type\s*[:\-]?\s*([A-Za-z][A-Za-z \-/]*?)(?=\s*(?:Total\s+Floor\s+Area|Assessment|Submission|Reference|Dwelling|\r|\n|$))",
        epr_text, re.IGNORECASE,
    )
    if m:
        cand = re.sub(r"\s{2,}", " ", m.group(1)).strip()
        if cand and any(t in cand.lower() for t in type_keywords):
            return cand

    # 2) Stacked layout: label column then value column.
    #    Find the "Property Type" label line, then take the value that
    #    sits just before the "Total Floor Area" value.
    lines = [l.strip() for l in epr_text.splitlines()]
    for i, line in enumerate(lines):
        if line.lower().startswith("property type"):
            tail = re.sub(r"(?i)property\s*type\s*[:\-]?\s*", "", line).strip()
            if tail:
                return re.sub(r"\s{2,}", " ", tail).strip()
            # Stacked: the value lives right above the Total Floor Area value
            # (which looks like "77 m2"). Walk forward to that area figure
            # and take the non-numeric line immediately before it.
            for j in range(i + 1, len(lines)):
                if re.match(r"^\d+(\.\d+)?\s*m", lines[j], re.IGNORECASE):
                    for k in range(j - 1, i, -1):
                        cand = lines[k].strip()
                        if cand and any(t in cand.lower() for t in type_keywords):
                            return re.sub(r"\s{2,}", " ", cand).strip()
                    break
            # Fallback: first dwelling-type-looking value line after label
            for nxt in lines[i + 1: i + 12]:
                if nxt and any(t in nxt.lower() for t in type_keywords):
                    return re.sub(r"\s{2,}", " ", nxt).strip()
            break

    # 3) Fallback: first line that looks like a dwelling type
    for line in lines:
        if any(k in line.lower() for k in type_keywords) and len(line) < 40:
            return re.sub(r"\s{2,}", " ", line).strip()

    return ""

# ================================================================
# SMARTSHEET
# ================================================================
def fetch_smartsheet_df(ss_token, sheet_id):
    url = f"https://api.smartsheet.eu/2.0/sheets/{sheet_id}"
    headers = {"Authorization": f"Bearer {ss_token}"}
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    sheet = r.json()
    col_map = {c["id"]: c["title"] for c in sheet["columns"]}
    rows = []
    for row in sheet["rows"]:
        d = {col_map[cell["columnId"]]: cell.get("displayValue", "") for cell in row["cells"]}
        rows.append(d)
    return pd.DataFrame(rows)

def fetch_client_spec_df(ss_token, second_sheet_id):
    url = f"https://api.smartsheet.eu/2.0/sheets/{second_sheet_id}"
    headers = {"Authorization": f"Bearer {ss_token}"}
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    sheet2 = resp.json()
    col_map2 = {c["id"]: c["title"] for c in sheet2["columns"]}
    rows2 = []
    for r2 in sheet2["rows"]:
        rows2.append({col_map2.get(c["columnId"], ""): c.get("displayValue", c.get("value", "")) for c in r2["cells"]})
    return pd.DataFrame(rows2)


# ================================================================
# PROPERTY MASTER LOOKUP USING PROPERTY ID
# ================================================================
PROPERTY_MASTER_ID_COL_OPTIONS = [
    "Property ID",
    "Property_ID",
    "Property Id",
    "property_id",
]

PROPERTY_MASTER_RISK_COLS = [
    CONSERVATION_AREA_COL,
    LISTED_BUILDING_COL,
    RADON_RISK_COL,
    FLOOD_RISK_COL,
    SUBS_2030_COL,
    SUBS_2070_COL,
    WDR_RISK_COL,
    RAINFALL_RISK_COL,
    WIND_SPEED_RISK_COL,
    WILDFIRE_RISK_COL,
]


def normalize_property_id(x):
    if pd.isna(x):
        return ""
    return re.sub(r"[^a-z0-9]", "", str(x).strip().lower())


def get_first_existing_col(df, possible_cols):
    for col in possible_cols:
        if col in df.columns:
            return col
    return None


def get_safe_value(row, col_name, default=""):
    if row is None or col_name not in row.index:
        return default
    val = row.get(col_name, default)
    if val is None or pd.isna(val):
        return default
    return str(val).strip()


def fetch_property_master_df(ss_token, property_master_sheet_id):
    return fetch_smartsheet_df(ss_token, property_master_sheet_id)


def find_property_master_row(df_master, property_id_value):
    if df_master is None or df_master.empty:
        return None

    master_id_col = get_first_existing_col(df_master, PROPERTY_MASTER_ID_COL_OPTIONS)
    if not master_id_col:
        raise ValueError(
            "Property Master sheet does not contain a Property ID column. "
            "Expected one of: Property ID, Property_ID, Property Id, property_id"
        )

    target_id = normalize_property_id(property_id_value)
    if not target_id:
        return None

    df_master = df_master.copy()
    df_master["_n_property_id"] = df_master[master_id_col].apply(normalize_property_id)
    matches = df_master[df_master["_n_property_id"] == target_id]

    if matches.empty:
        return None

    return matches.iloc[0]

# ================================================================
# COORDINATES
# ================================================================
def dms_to_dd(deg, minutes, sec, direction):
    dd = float(deg) + float(minutes) / 60 + float(sec) / 3600
    return round(-dd if direction in ["S", "W"] else dd, 5)

def extract_lat_lon_from_cr(path):
    text = read_pdf_text(path)
    lat_match = re.search(r"Latitude[: ]\s*(\d+)[° ]\s*(\d+)'?\s*(\d+)''?\s*([NSEW])", text)
    lon_match = re.search(r"Longitude[: ]\s*(\d+)[° ]\s*(\d+)'?\s*(\d+)''?\s*([NSEW])", text)
    lat = dms_to_dd(*lat_match.groups()) if lat_match else ""
    lon = dms_to_dd(*lon_match.groups()) if lon_match else ""
    return lat, lon

# ================================================================
# EXTRACTION HELPERS
# ================================================================
def safe_search(pattern, text, flags=0):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else ""

def extract_date_built_from_sn(path):
    lines = read_pdf_page_lines(path, 0)
    for i, line in enumerate(lines):
        if "Date Built" in line:
            for l in lines[i: i + 6]:
                m = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", l)
                if m: return f"{m.group(1)}-{m.group(2)}"
            m2 = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", read_pdf_text(path))
            if m2: return f"{m2.group(1)} - {m2.group(2)}"
            break
    return ""

def extract_num_storeys_from_sn(path):
    lines = read_pdf_page_lines(path, 0)
    for i, line in enumerate(lines):
        if "Number of" in line and "Storeys" in " ".join(lines[i: i + 3]):
            for j in range(i + 1, min(i + 6, len(lines))):
                m = re.search(r"\b(\d+)\b", lines[j])
                if m: return m.group(1)
            break
    return ""

def extract_exposure_zone_from_cr(path):
    for line in read_pdf_page_lines(path, 2):
        if "Exposure" in line and "Zone" in line:
            m = re.search(r"Zone\s*([A-Za-z]+)", line)
            if m: return m.group(1).strip()
    return ""

def extract_vulnerable_flag_from_cr(path):
    with fitz.open(path) as doc:
        last_page = len(doc) - 1
    lines = read_pdf_page_lines(path, last_page)
    questions = [
        "How many are children under the age of 18?",
        "How many of a pensionable age?",
        "How many with disabilities?"
    ]
    values = []
    for line in lines:
        s = line.strip()
        for q in questions:
            if s.startswith(q):
                part = s.split("?", 1)[1].strip()
                values.append(int(part) if part.isdigit() else 0)
    return "Yes" if any(v > 0 for v in values) else "No"

def extract_constraints_and_construction_from_cr(path):
    text = "\n".join(read_pdf_page_lines(path, 2))
    m_con  = re.search(r"Property\s+Constraints\s*([^\n\r]+)", text)
    m_cons = re.search(r"Property\s+Construction\s*([^\n\r]+)", text)
    constraints_val   = m_con.group(1).strip() if m_con else ""
    construction_flag = "Yes" if m_cons and m_cons.group(1).strip().lower().startswith("traditional") else "No"
    return constraints_val, construction_flag

def get_building_orientation_direction(path):
    text = "\n".join(read_pdf_page_lines(path, 2))
    m = re.search(r"Building\s+Orientation\s*\(degrees\)\s*([0-9]+(?:\.[0-9]+)?)", text)
    if not m: return ""
    angle = float(m.group(1))
    table = [
        (15, "North"), (45, "North East"), (75, "East"), (105, "East"),
        (135, "South East"), (165, "South"), (195, "South"), (225, "South West"),
        (255, "West"), (285, "West"), (315, "North West"), (345, "North"), (360, "North")
    ]
    for heading, direction in table:
        if angle <= heading: return direction
    return ""

def extract_highest_bedroom_from_contents(cr_page2_lines):
    highest_bedroom = 0
    found_plain_bedroom = False
    for line in cr_page2_lines:
        m = re.search(r"\bBedroom\s+(\d+)\b", line.strip(), flags=re.IGNORECASE)
        if m:
            highest_bedroom = max(highest_bedroom, int(m.group(1)))
        elif re.search(r"\bBedroom\b", line.strip(), flags=re.IGNORECASE):
            found_plain_bedroom = True
    if highest_bedroom == 0 and found_plain_bedroom:
        highest_bedroom = 1
    return highest_bedroom

# ================================================================
# ROOM EXTRACTION FROM CONTENTS PAGE  -- FIXED
# Now returns BOTH cleaned room names AND raw contents lines,
# so the downstream page-range lookup matches the working
# original_script.txt behaviour.
# ================================================================
_PAGE_NOISE_PATTERN = re.compile(r"^page\s+\d+(\s+of\s+\d+)?$", re.IGNORECASE)

def extract_rooms_from_contents(cr_pdf_path):
    """
    Returns a tuple: (rooms_list, raw_contents_lines)
      - rooms_list: cleaned, ordered list of room name strings
      - raw_contents_lines: page-1 lines (whitespace-stripped, blanks removed)
        used later by get_room_page_range() for accurate page lookup.
    """
    raw_lines = read_pdf_page_lines(cr_pdf_path, 1)

    # Keep raw contents (stripped, no blanks) for downstream page-range lookup
    contents_raw = [l.strip() for l in raw_lines if l.strip()]

    # Build cleaned name list (drop pure digits + page noise + header rows)
    cleaned = []
    for x in contents_raw:
        if x.isdigit():
            continue
        if _PAGE_NOISE_PATTERN.match(x.lower()):
            continue
        if x.lower() in ("contents", "property details"):
            continue
        cleaned.append(x)

    # Find start of room block
    possible_starts = [
        "Kitchen", "Kitchen & storage", "Kitchen and Storage",
        "Living Room", "Lounge", "Dining Room", "Dining",
        "Bedroom", "Bedroom 1", "Hall", "Hallway",
        "Bathroom", "WC", "w/c", "Toilet", "Utility"
    ]
    start_idx = 0
    for i, item in enumerate(cleaned):
        if any(s.lower() in item.lower() for s in possible_starts):
            start_idx = i
            break

    end_markers = ["occupancy", "additional", "summary",
                   "recommendations", "annex", "appendix"]
    end_idx = next(
        (i for i, item in enumerate(cleaned)
         if any(m in item.lower() for m in end_markers)),
        len(cleaned)
    )

    rooms = cleaned[start_idx:end_idx]
    non_room_keywords = ["elevation", "floorplan", "floor plan", "summary",
                         "additional information", "annex", "appendix", "notes"]
    rooms = [r for r in rooms
             if not any(kw in r.lower() for kw in non_room_keywords)]

    # Deduplicate while preserving order
    seen, dedup = set(), []
    for r in rooms:
        key = r.strip().lower()
        if key not in seen:
            seen.add(key)
            dedup.append(r)

    return dedup, contents_raw


# ================================================================
# ROOM EXTRACTION QUESTIONS
# ================================================================
Q_CONDENSATION = "Is there any evidence of condensation, damp or mould growth in this room?"
Q_BG_AREA      = "Background Ventilation Area (mm2)"
Q_TRICKLE      = "Do they have trickle vents?"
Q_UNDERCUTS    = "Are the undercuts on all the internal doors"
Q_WINDOWS      = "Does the room have any windows?"
Q_FANS         = "Does the room have fans fitted?"

def extract_yes_no_strict(lines, question):
    for line in lines:
        if line.startswith(question):
            parts = line.split("?", 1)
            if len(parts) == 2 and parts[1].strip() in ("Yes", "No"):
                return parts[1].strip()
            return None
    return None

def extract_mm2_strict(lines):
    for line in lines:
        if line.startswith(Q_BG_AREA):
            tail = line.replace(Q_BG_AREA, "").strip()
            return tail if tail.isdigit() else None
    return None

def clean_lines(lines):
    return [l.strip() for l in lines if l.strip() and not _PAGE_NOISE_PATTERN.match(l.strip().lower())]


# ================================================================
# PAGE-RANGE LOOKUP -- matches original_script.txt behaviour
# Scans raw contents list forward from the room name to find
# the next digit (= start page) and the digit after that
# (= start page of next room).
# ================================================================
def get_room_page_range(contents_raw, room):
    try:
        idx = contents_raw.index(room)
    except ValueError:
        return None, None

    start_page = None
    i = None
    for i in range(idx + 1, len(contents_raw)):
        if str(contents_raw[i]).strip().isdigit():
            start_page = int(str(contents_raw[i]).strip())
            break

    if start_page is None:
        return None, None

    next_page = None
    for j in range(i + 1, len(contents_raw)):
        if str(contents_raw[j]).strip().isdigit():
            next_page = int(str(contents_raw[j]).strip())
            break

    if next_page is None:
        return start_page - 1, start_page + 2

    return start_page - 1, next_page - 2


# ================================================================
# ROOM DATA EXTRACTION  -- FIXED
# Uses the raw contents list + get_room_page_range() to find
# the exact page range for each room. This is what makes the
# question-answer extraction reliable.
# ================================================================
def extract_room_data(cr_pdf_path, contents_raw, room):
    empty = {"condensation": None, "background": None, "trickle": None,
             "undercuts": None, "windows": None, "fans": None}

    start_idx, end_idx = get_room_page_range(contents_raw, room)
    if start_idx is None:
        return empty

    all_lines = []
    try:
        with fitz.open(cr_pdf_path) as doc:
            total_pages = len(doc)
        safe_start = max(0, start_idx)
        safe_end   = min(end_idx + 1, total_pages)
        for p in range(safe_start, safe_end):
            all_lines.extend(read_pdf_page_lines(cr_pdf_path, p))
    except Exception:
        return empty

    lines = clean_lines(all_lines)
    return {
        "condensation": extract_yes_no_strict(lines, Q_CONDENSATION),
        "background":   extract_mm2_strict(lines),
        "trickle":      extract_yes_no_strict(lines, Q_TRICKLE),
        "undercuts":    extract_yes_no_strict(lines, Q_UNDERCUTS),
        "windows":      extract_yes_no_strict(lines, Q_WINDOWS),
        "fans":         extract_yes_no_strict(lines, Q_FANS),
    }


# ================================================================
# FLOW RATE HELPERS
# ================================================================
def get_min_lps_for_room(room_name):
    name = room_name.lower()
    if "kitchen" in name:                                      return "60 l/s minimum"
    if "bathroom" in name or "wetroom" in name:                return "15 l/s minimum"
    if "wc" in name or "toilet" in name or "w/c" in name:     return "6 l/s minimum"
    if "utility" in name:                                      return "30 l/s minimum"
    return "N/A"

def get_min_lps_for_room_q(room_name):
    name = room_name.lower()
    if "kitchen" in name:                                      return "13 l/s minimum"
    if "bathroom" in name or "wetroom" in name:                return "8 l/s minimum"
    if "wc" in name or "toilet" in name or "w/c" in name:     return "6 l/s minimum"
    if "utility" in name:                                      return "6 l/s minimum"
    return "N/A"

def extract_lps_number(value):
    if not value: return 0
    if isinstance(value, str) and "l/s" in value:
        try:    return int(value.split()[0])
        except: return 0
    return 0

# ================================================================
# EXCEL HELPERS
# ================================================================
def find_row_by_label(ws, label, col="A"):
    for r in range(1, ws.max_row + 1):
        val = ws[f"{col}{r}"].value
        if isinstance(val, str) and val.strip() == label:
            return r
    raise ValueError(f"Row with label '{label}' not found")

def parse_area_m2(value):
    if value is None: return 0.0
    if isinstance(value, (int, float)): return float(value)
    if isinstance(value, str):
        num = "".join(ch for ch in value if ch.isdigit() or ch == ".")
        return float(num) if num else 0.0
    return 0.0

def get_master_cell(ws, cell_ref):
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            return merged.coord.split(":")[0]
    return cell_ref

def unmerge_in_range(ws, start_row):
    affected = []
    for m in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = range_boundaries(str(m))
        if r1 >= start_row:
            affected.append((c1, r1, c2, r2))
            ws.unmerge_cells(str(m))
    return affected

def remerge(ws, ranges):
    for c1, r1, c2, r2 in ranges:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# ================================================================
# WRITE ROOM ROW
#
# Changes from previous version:
#   • Point 3 — WC / W/C / Utility:  B = CR (condensation) value,
#       D = "No minimum".
#   • Point 2 — Bathroom & Kitchen ONLY:  column M (Proposed) is driven
#       by the Background Ventilation Area (D). Value present -> "Sealed";
#       "No"/"No minimum"/blank -> "No Works". (Other wet rooms keep the
#       fans-based logic.)
#   • Point 4 — Hallway:  column N is always "No Works".
# ================================================================
def write_room_row(ws, row, d, room):
    def v(x): return x if x is not None else "Not Extracted"
    room_lower = room.lower()
    bg_value   = d["background"]
    fans_value = v(d["fans"])
    wet_room_keys = ["kitchen", "bathroom", "wetroom", "wet room", "wc", "w/c", "toilet", "utility"]
    is_wet  = any(w in room_lower for w in wet_room_keys)
    is_wcu  = is_wc_or_utility(room)
    is_bk   = is_bathroom_or_kitchen(room)
    is_hall = "hall" in room_lower

    # --- Column B: always the CR (condensation) value, incl. WC/W-C/Utility ---
    ws[f"B{row}"].value = v(d["condensation"])

    # --- Column C: background ventilation requirement ---
    if bg_value is None:
        ws[f"C{row}"].value = "No"
    elif "bathroom" in room_lower or "wetroom" in room_lower:
        ws[f"C{row}"].value = "4000mm2"
    else:
        ws[f"C{row}"].value = "8000mm2"

    # --- Column D: "No minimum" for WC/W-C/Utility, else the background area ---
    if is_wcu:
        ws[f"D{row}"].value = "No minimum"
    else:
        ws[f"D{row}"].value = "No" if bg_value is None else f"{bg_value}mm2"

    ws[f"E{row}"].value = v(d["undercuts"])
    ws[f"F{row}"].value = v(d["windows"])

    if is_wet:
        ws[f"G{row}"].value = "IEV Present" if fans_value == "Yes" else "None"
    else:
        ws[f"G{row}"].value = "N/A"

    ws[f"H{row}"].value = get_min_lps_for_room(room)

    cond_val = d["condensation"]
    ws[f"K{row}"].value = (
        "No" if cond_val == "No" else
        "Clean" if cond_val == "Yes" else
        "Not Extracted"
    )
    ws[f"L{row}"].value = "N/A" if is_wet else "4000mm2"

    col_c_val     = ws[f"C{row}"].value
    undercuts_val = ws[f"E{row}"].value

    # --- Column M (Proposed) ---
    if is_bk:
        # Bathroom & Kitchen: driven by Background Ventilation Area (D).
        # Value present -> Sealed ; "No"/"No minimum"/blank -> No Works.
        ws[f"M{row}"].value = "Sealed" if bg_value is not None else "No Works"
    elif is_wet:
        ws[f"M{row}"].value = "Sealed" if fans_value == "Yes" else "No Works"
    else:
        ws[f"M{row}"].value = "Required" if col_c_val == "No" and undercuts_val == "Yes" else "No Works"

    # --- Column N: Hallway is always "No Works" ---
    if is_hall:
        ws[f"N{row}"].value = "No Works"
    else:
        ws[f"N{row}"].value = (
            "Required" if d["undercuts"] == "No" else
            "No Works" if d["undercuts"] == "Yes" else
            "Not Extracted"
        )

    ws[f"O{row}"].value = (
        "Required" if d["windows"] == "No" else
        "No Works" if d["windows"] == "Yes" else
        "Not Extracted"
    )

    if is_wet:
        ws[f"P{row}"].value = "Upgrade dMEV" if fans_value == "Yes" else "Install dMEV"
    else:
        ws[f"P{row}"].value = "N/A"

    ws[f"Q{row}"].value = get_min_lps_for_room_q(room)
    ws[f"R{row}"].value = "Information Unavailable"

# ================================================================
# MAIN ROOM INSERTION (Sheet 11.0)  -- FIXED
# Now uses (room_names, contents_raw) from extract_rooms_from_contents
# and passes contents_raw into extract_room_data for accurate
# page-range lookup.
# ================================================================
def insert_rooms_with_template(cr_pdf_path, workbook, sheet_name="11.0",
                                insert_at_row=5, template_row=4,
                                room_col_left="A", room_col_right="J",
                                tfa=None, highest_bedroom=0, log=None):
    def _log(msg):
        if log is not None: log(msg)

    ws = workbook[sheet_name]

    room_names, contents_raw = extract_rooms_from_contents(cr_pdf_path)
    _log(f"📋 Rooms extracted ({len(room_names)}): {room_names}")

    if not room_names:
        raise ValueError("No rooms found in Contents page.")

    room_names = sorted(room_names, key=wet_room_sort_key)
    _log(f"📋 Rooms after sort: {room_names}")

    original_heights = {r: d.height for r, d in ws.row_dimensions.items() if d.height is not None}
    affected_merges  = unmerge_in_range(ws, insert_at_row)
    ws.insert_rows(insert_at_row, amount=len(room_names))

    for r, h in original_heights.items():
        ws.row_dimensions[r + len(room_names) if r >= insert_at_row else r].height = h

    tpl_height = ws.row_dimensions[template_row].height
    for i in range(len(room_names)):
        ws.row_dimensions[insert_at_row + i].height = tpl_height
        for c in range(1, ws.max_column + 1):
            ws.cell(insert_at_row + i, c)._style = copy(ws.cell(template_row, c)._style)

    for i, room in enumerate(room_names):
        row  = insert_at_row + i
        ws[f"{room_col_left}{row}"].value = room
        ws[f"{room_col_right}{row}"].value = room
        data = extract_room_data(cr_pdf_path, contents_raw, room)
        write_room_row(ws, row, data, room)
        found = " ".join(f"{k}={v}" for k, v in data.items() if v is not None)
        _log(f"  ✓ {room}: {found or 'No questions found'}")

    ws4       = workbook["4.0"]
    room_start = insert_at_row
    room_end   = insert_at_row + len(room_names) - 1

    ws4["H5"].value = (
        "Yes" if any(ws[f"P{r}"].value == "Upgrade dMEV" for r in range(room_start, room_end + 1))
        else "No"
    )
    ws4["B8"].value = (
        "Yes" if any(ws[f"B{r}"].value == "Yes" for r in range(room_start, room_end + 1))
        else "No"
    )

    total_lps_h = sum(extract_lps_number(ws[f"H{insert_at_row + i}"].value) for i in range(len(room_names)))
    total_lps_q = sum(extract_lps_number(ws[f"Q{insert_at_row + i}"].value) for i in range(len(room_names)))
    total_row   = insert_at_row + len(room_names) + 1
    ws[f"H{total_row}"].value = f"{total_lps_h} l/s"
    ws[f"Q{total_row}"].value = f"{total_lps_q} l/s"

    floor_area_row = find_row_by_label(ws, "Whole building ventilation rate (l/s) - floor area (m2)", col="A")
    ws[f"G{floor_area_row}"].value = tfa
    ws[f"P{floor_area_row}"].value = tfa
    floor_rate = custom_round(parse_area_m2(tfa) * 0.3)
    ws[f"H{floor_area_row}"].value = f"{floor_rate} l/s"
    ws[f"Q{floor_area_row}"].value = f"{floor_rate} l/s"

    bedroom_row  = find_row_by_label(ws, "Whole building ventilation rate (l/s) - bedrooms", col="A")
    ws[f"G{bedroom_row}"].value = f"{highest_bedroom} bed"
    ws[f"P{bedroom_row}"].value = f"{highest_bedroom} bed"
    bed_rate_raw = {1: 19, 2: 25, 3: 31, 4: 37, 5: 43}.get(highest_bedroom, 0)
    bed_rate     = custom_round(bed_rate_raw)
    ws[f"H{bedroom_row}"].value = f"{bed_rate} l/s"
    ws[f"Q{bedroom_row}"].value = f"{bed_rate} l/s"

    final_row    = find_row_by_label(ws, "Whole building ventilation rate - final", col="A")
    final_rate_h = custom_round(max(floor_rate, bed_rate, total_lps_h))
    final_rate_q = custom_round(max(floor_rate, bed_rate, total_lps_q))
    ws[f"H{final_row}"].value = f"{final_rate_h} l/s"
    ws[f"Q{final_row}"].value = f"{final_rate_q} l/s"

    remerge(ws, [(c1, r1 + len(room_names), c2, r2 + len(room_names)) for c1, r1, c2, r2 in affected_merges])
    return room_names

# ================================================================
# WRITE SHEET 4.0
# ================================================================
def write_sheet_4_base(ws, wb_data):
    d = wb_data
    ws[get_master_cell(ws, "B3")]  = d["dwelling_address"]
    ws[get_master_cell(ws, "B4")]  = d["property_id_value"]
    ws[get_master_cell(ws, "B5")]  = d["local_authority_name"]   # Local Authority -> 4.0 B5 (only)
    ws[get_master_cell(ws, "B6")]  = d["sap_current"]
    ws[get_master_cell(ws, "E3")]  = d["sap_current"]
    ws[get_master_cell(ws, "E9")]  = d["sap_potential"]
    ws[get_master_cell(ws, "B7")]  = d["tenure"]
    ws[get_master_cell(ws, "B14")] = d["date_built"]
    ws[get_master_cell(ws, "B15")] = d["wall_type"]
    ws[get_master_cell(ws, "B16")] = d["property_type"]
    ws[get_master_cell(ws, "B17")] = d["num_storeys"]
    ws[get_master_cell(ws, "B18")] = d["highest_bedroom"] if d["highest_bedroom"] > 0 else ""
    ws[get_master_cell(ws, "B19")] = d["exposure_zone"]
    ws[get_master_cell(ws, "B20")] = d["property_construction_flag"]
    ws[get_master_cell(ws, "B21")] = d["conservation_flag"]
    ws[get_master_cell(ws, "B22")] = "No"
    ws[get_master_cell(ws, "B24")] = d["property_constraints_val"]
    ws[get_master_cell(ws, "B25")] = d["vulnerable_flag"]
    ws[get_master_cell(ws, "E4")]  = d["tfa_sqm"]
    ws[get_master_cell(ws, "E10")] = d["tfa_sqm"]
    ws[get_master_cell(ws, "E5")]  = d["extra1_last"]
    ws[get_master_cell(ws, "E6")]  = d["extra1"]
    ws[get_master_cell(ws, "E11")] = d["extra2_last"]
    ws[get_master_cell(ws, "E12")] = d["extra2"]
    conservation = d.get("conservation_flag", "No")
    if str(conservation).strip().lower() == "yes":
        for cell in ["B10", "B12"]: ws[get_master_cell(ws, cell)] = "TBC"
        for cell in ["B11", "B13"]: ws[get_master_cell(ws, cell)] = "Pending"
    else:
        for cell in ["B10", "B11", "B12", "B13"]: ws[get_master_cell(ws, cell)] = "No"

def write_sheet_4_climate_risk(ws, wb_data):
    d = wb_data
    ws[get_master_cell(ws, "E16")] = d["radon_risk"]
    ws[get_master_cell(ws, "E17")] = d["flood_risk"]
    ws[get_master_cell(ws, "E18")] = d["subs2030"]
    ws[get_master_cell(ws, "E19")] = d["subs2070"]
    ws[get_master_cell(ws, "E20")] = d["wdr_risk"]
    ws[get_master_cell(ws, "E21")] = d["rainfall_risk"]
    ws[get_master_cell(ws, "E22")] = d["wind_speed_risk"]
    ws[get_master_cell(ws, "E23")] = d["wildfire_risk"]

# ================================================================
# WRITE SHEET 10.0
#
# Change: Local Authority no longer goes here. Instead:
#   • B3 <- Project Name (new input)
#   • F3 <- REF
#   • B3 is explicitly cleared (left empty).
# ================================================================
def write_sheet_10(workbook, project_name, f3_value, log=None):
    if "10.0" not in workbook.sheetnames:
        if log: log("⚠️  Sheet '10.0' not found — skipping.")
        return
    ws = workbook["10.0"]
    ws[get_master_cell(ws, "B3")] = None          # was Local Authority — now empty
    ws[get_master_cell(ws, "B3")] = project_name  # Project Name
    ws[get_master_cell(ws, "F3")] = f3_value      # REF
    if log: log(f"  ✓ Sheet 10.0: B5(Project)='{project_name}', F3(REF)='{f3_value}', B3=cleared")

# ================================================================
# WRITE SHEET 7.0  -- REWRITTEN
#
# spec_col is a SINGLE Smartsheet column name. Each output
# (Prelim / Construction) passes its own column.
#
# Routing rules:
#   • Standard measures      -> their fixed template rows.
#   • EXTRA measures (anything not in the standard list):
#       - heating / renewable  -> next EMPTY renewable row (17-21) NOT used
#                                 by a selected standard measure.
#       - fabric / ventilation -> the "Other please specify" row (row 14).
#   • An extra measure NEVER overwrites a selected standard measure's row.
#   • Every extra measure is ALWAYS Category = "Minor" (column H).
#   • Sequence (col G): standard measures get 1..N install order; extras
#     continue N+1, N+2… (renewable extras first, then fabric extras).
#   • Communal PV is handled specially (fills SPV row, relabels C22).
# ================================================================
def write_sheet_7(ws, wb_data, df_specs, selected_measures, orientation_direction,
                  ws4, spec_col=None, log=None):
    def _log(msg):
        if log: log(msg)

    measure_to_cell = {
        "Trickle Vents":       "A2",
        "dMEV":                "A3",
        "Door undercuts":      "A4",
        "LI (B9)":             "A5",
        "CWI":                 "A6",
        "EWI":                 "A7",
        "IWI":                 "A8",
        "UFI":                 "A9",
        "FRI":                 "A10",
        "RiR and/or Skeiling": "A11",
        "Windows":             "A12",
        "Doors":               "A13",
        # row 14 = "Other please specify" (reserved for extra fabric measures)
        "Boiler/FTCH":         "A17",
        "ESH":                 "A18",
        "ASHP":                "A19",
        "GSHP":                "A20",
        "Solar Thermal":       "A21",
        "SPV":                 "A22",
    }

    abbrev_map = {
        "Trickle Vents":       "TV",
        "dMEV":                "dMEV",
        "Door undercuts":      "DRU",
        "LI (B9)":             "LI",
        "CWI":                 "CWI",
        "EWI":                 "EWI",
        "IWI":                 "IWI",
        "UFI":                 "UFI",
        "FRI":                 "FRI",
        "RiR and/or Skeiling": "RiR/SKI",
        "Windows":             "Windows",
        "Doors":               "Doors",
        "Boiler/FTCH":         "BLR",
        "ESH":                 "ESH",
        "ASHP":                "ASHP",
        "GSHP":                "GSHP",
        "Solar Thermal":       "ST",
        "SPV":                 "SPV",
        # ---- extra measures (drawing keys for column C) ----
        "Sub Floor Ventilation":  "AirEx Floorvent",
        "Sub-Floor Ventilation":  "AirEx Floorvent",
        "AirEx Floorvent":        "AirEx Floorvent",
        "AirEx":                  "AirEx Floorvent",
        "Air Brick":              "AirEx Floorvent",
        "Electric Meter Dual":    "EM-D",
        "Solar Diverter":         "SD",
        "Solar iBoost":           "SD",
        "LEL":                    "LEL",
        "HWC":                    "HWC",
        "Draught-Proofing":       "DPRF",
        "Upgrade boiler, same fuel":                "UBLR",
        "Heat Recovery System for Mixer Showers":   "HRSMS",
        "Communal PV":            "SPV",
    }

    SEQUENCE_ORDER = [
        ["dMEV"], ["CWI"], ["EWI"], ["IWI"], ["LI (B9)"], ["FRI"],
        ["RiR and/or Skeiling"], ["Windows", "Doors"], ["Boiler/FTCH"],
        ["ESH"], ["ASHP"], ["GSHP"], ["Solar Thermal"], ["HWC"], ["LEL"], ["SPV"],
    ]

    HEATING_RENEWABLE_KEYWORDS = [
        "boiler", "ftch", "esh", "storage heater", "ashp", "gshp",
        "heat pump", "solar", "pv", "photovolta", "thermal", "diverter",
        "immersion", "battery", "meter", "hwc", "cylinder", "iboost",
        "renewable", "lel",
    ]

    def is_heating_renewable(name):
        n = name.lower()
        return any(k in n for k in HEATING_RENEWABLE_KEYWORDS)

    def get_spec(abbr):
        # Read the spec value from the single chosen column only.
        if df_specs.empty or "Abbreviation" not in df_specs.columns:
            return ""
        if not spec_col or spec_col not in df_specs.columns:
            return ""
        spec_row = df_specs[df_specs["Abbreviation"] == abbr]
        if spec_row.empty:
            return ""
        val = spec_row.iloc[0].get(spec_col, "")
        return str(val).strip() if val and str(val).strip() else ""

    # Candidate names for the "Description" column on the Client Spec sheet.
    DESC_COL_CANDIDATES = ["Description", "Measure Description", "Measure",
                           "Full Description", "Measure Name"]

    def _spec_desc_col():
        for c in DESC_COL_CANDIDATES:
            if c in df_specs.columns:
                return c
        return None

    def lookup_extra_from_spec(token):
        """Match an EXTRA measure (IOE token) against the Client Spec sheet by
        Abbreviation OR Description, allowing PARTIAL / half-word matches
        (e.g. IOE 'AirEx' matches Abbreviation 'AirEx Floorvent').

        Returns (description_for_B, abbreviation_for_C, spec_for_D), or
        None if no match is found — in which case the caller IGNORES the
        measure entirely (writes nothing)."""
        token = str(token).strip()
        if not token or df_specs.empty:
            return None

        norm = lambda x: re.sub(r"\s+", " ", str(x).strip().lower())
        nt = norm(token)
        if not nt:
            return None

        desc_col = _spec_desc_col()
        has_abbr = "Abbreviation" in df_specs.columns

        # Rank candidates (lower = better):
        #   0 exact Abbreviation       1 exact Description
        #   2 token inside Abbreviation 3 Abbreviation inside token
        #   4 token inside Description  5 Description inside token
        # Substring (ranks 2-5) only when the token is >= 3 chars to avoid
        # tiny accidental matches.
        best, best_rank = None, 99
        allow_sub = len(nt) >= 3
        for _, r in df_specs.iterrows():
            na = norm(r.get("Abbreviation", "")) if has_abbr else ""
            nd = norm(r.get(desc_col, "")) if desc_col else ""
            rank = None
            if has_abbr and na and na == nt:
                rank = 0
            elif desc_col and nd and nd == nt:
                rank = 1
            elif allow_sub and has_abbr and na and len(na) >= 3 and nt in na:
                rank = 2
            elif allow_sub and has_abbr and na and len(na) >= 3 and na in nt:
                rank = 3
            elif allow_sub and desc_col and nd and len(nd) >= 3 and nt in nd:
                rank = 4
            elif allow_sub and desc_col and nd and len(nd) >= 3 and nd in nt:
                rank = 5
            if rank is not None and rank < best_rank:
                best, best_rank = r, rank
                if rank == 0:
                    break

        if best is None:
            return None  # not in Client Spec → ignore the measure

        abbr = str(best.get("Abbreviation", "")).strip() if has_abbr else ""
        desc = str(best.get(desc_col, "")).strip() if desc_col else ""
        spec = ""
        if spec_col and spec_col in df_specs.columns:
            v = best.get(spec_col, "")
            spec = str(v).strip() if v and str(v).strip() else ""
        return (desc or abbr or token), (abbr or desc or token), spec

    def set_cell_unmerged(cell_ref, value):
        col_letter = "".join(ch for ch in cell_ref if ch.isalpha())
        row_num    = int("".join(ch for ch in cell_ref if ch.isdigit()))
        col_idx    = column_index_from_string(col_letter)
        for merged in list(ws.merged_cells.ranges):
            c1, r1, c2, r2 = range_boundaries(str(merged))
            if r1 <= row_num <= r2 and c1 <= col_idx <= c2:
                ws.unmerge_cells(str(merged))
                break
        ws[get_master_cell(ws, cell_ref)] = value

    def find_other_specify_row():
        for r in range(2, 16):
            b = ws[get_master_cell(ws, f"B{r}")].value
            if isinstance(b, str) and "other" in b.lower() and "specify" in b.lower():
                return r
        return 14  # fallback to known template position

    # ===========================================================
    # 1) Place STANDARD measures + handle Communal PV specially
    # ===========================================================
    measures_list  = [m.strip() for m in str(selected_measures).split(",") if m.strip()]
    known_selected = set()
    used_rows      = set()
    extra_measures = []
    spv_included   = False

    for measure in measures_list:
        # --- Communal PV: fills the SPV row but relabels C22 ---
        if "communal pv" in measure.lower():
            spv_included = True
            row_num = int(measure_to_cell["SPV"][1:])     # 22
            ws[get_master_cell(ws, f"D{row_num}")] = get_spec("SPV")
            set_cell_unmerged("C22", "Communal PV")
            known_selected.add("SPV")
            used_rows.add(row_num)
            _log(f"  ✓ Communal PV → row {row_num} (SPV spec), C22='Communal PV'")
            continue

        matched_key = None
        for key in measure_to_cell:
            if key.lower() in measure.lower():
                matched_key = key
                break

        if matched_key is not None:
            if matched_key == "SPV":
                spv_included = True
            row_num = int(measure_to_cell[matched_key][1:])
            ws[get_master_cell(ws, f"D{row_num}")] = get_spec(abbrev_map.get(matched_key, ""))
            known_selected.add(matched_key)
            used_rows.add(row_num)
            _log(f"  ✓ Standard measure '{matched_key}' → row {row_num}")
        else:
            extra_measures.append(measure)

    # ===========================================================
    # 2) Place EXTRA measures
    # ===========================================================
    renewable_slots      = list(range(17, 22))   # 17-21 reusable; 22-23 = SPV
    renewable_extra_rows = []
    fabric_extra_rows    = []

    def next_empty_renewable_row():
        for r in renewable_slots:
            if r in used_rows:
                continue
            if ws[get_master_cell(ws, f"D{r}")].value in (None, ""):
                return r
        return None

    other_row  = find_other_specify_row()
    other_used = False

    for measure in extra_measures:
        # Pull from the Client Spec Smartsheet (NOT from IOE / Property Master):
        #   B <- Description, C <- Abbreviation, D <- spec column.
        # If the measure is not found in the Client Spec sheet → IGNORE it.
        found = lookup_extra_from_spec(measure)
        if found is None:
            _log(f"  ⛔ Extra '{measure}' not found in Client Spec "
                 f"(no Abbreviation/Description match) — ignored")
            continue
        desc, abbr, spec = found
        classify_text = f"{measure} {desc} {abbr}"

        if is_heating_renewable(classify_text):
            target = next_empty_renewable_row()
            if target is None:
                _log(f"  ⚠️  No empty renewable slot for extra '{measure}' — skipped")
                continue
            renewable_extra_rows.append(target)
            zone = "renewable table"
        else:
            if other_used:
                _log(f"  ⚠️  Extra fabric measure '{measure}' skipped — "
                     f"'Other please specify' row already used")
                continue
            target = other_row
            other_used = True
            fabric_extra_rows.append(target)
            zone = "Other please specify"

        set_cell_unmerged(f"B{target}", desc)               # Description (Client Spec sheet)
        set_cell_unmerged(f"C{target}", abbr)               # Abbreviation (Client Spec sheet)
        ws[get_master_cell(ws, f"D{target}")] = spec        # Specification (Client Spec sheet)
        ws[get_master_cell(ws, f"H{target}")] = "Minor"     # ALWAYS Minor
        used_rows.add(target)
        _log(f"  ➕ Extra '{measure}' → row {target} [{zone}] "
             f"B='{desc}' C='{abbr}', Category=Minor")

    # ===========================================================
    # 3) Solar PV orientation / shading
    # ===========================================================
    if spv_included:
        if orientation_direction:
            ws[get_master_cell(ws, "F22")] = orientation_direction
        ws[get_master_cell(ws, "F23")] = "None or little"
        _log(f"  ✓ PV included → F22='{orientation_direction}', F23='None or little'")
    else:
        ws[get_master_cell(ws, "F22")] = None
        ws[get_master_cell(ws, "F23")] = None

    # ===========================================================
    # 4) Major-works flags on sheet 4.0 (H3/H4)
    # ===========================================================
    major_fabric_keys = {"LI (B9)", "CWI", "EWI", "IWI", "FRI", "Windows", "Doors"}
    condition_met = any(k in known_selected for k in major_fabric_keys)
    ws4["H3"].value = "Yes" if condition_met else "No"
    ws4["H4"].value = "Yes" if condition_met else "No"

    # ===========================================================
    # 5) Installation sequence (column G)
    # ===========================================================
    seq_number = 1
    for group in SEQUENCE_ORDER:
        if any(name in known_selected for name in group):
            for name in group:
                if name in known_selected and name in measure_to_cell:
                    row_num = int(measure_to_cell[name][1:])
                    if row_num in renewable_extra_rows or row_num in fabric_extra_rows:
                        continue
                    ws[get_master_cell(ws, f"G{row_num}")] = seq_number
            seq_number += 1

    for extra_row in renewable_extra_rows + fabric_extra_rows:
        ws[get_master_cell(ws, f"G{extra_row}")] = seq_number
        seq_number += 1

# ================================================================
# CORE PROCESSING FUNCTION
# prelim_spec_col  -> column used for Prelim output's Sheet 7.0
# construction_spec_col -> column used for Construction output's Sheet 7.0
# ================================================================
def process_property(
    cr_path, sn_path, epr_path,
    template_bytes,
    ss_token, sheet_id, property_master_sheet_id, second_sheet_id,
    local_authority_name, project_name, sheet10_f3,
    prelim_spec_col, construction_spec_col,
    log
):
    start_time = time.time()

    # -- Read PDFs ---
    sn_text  = read_pdf_text(sn_path)
    epr_text = read_pdf_text(epr_path)

    # -- Address matching --
    dwelling_address = extract_property_address_from_cr(cr_path)
    log(f"📍 Extracted address: {dwelling_address}")

    df1 = pd.DataFrame({"Address": [dwelling_address]})
    df1["Postcode"]      = df1["Address"].apply(extract_postcode)
    df1["house_number"]  = df1["Address"].apply(extract_house_number)
    df1["street_number"] = df1.apply(lambda r: extract_street_number(r["Address"], r["house_number"]), axis=1)
    df1["n_postcode"]    = df1["Postcode"].apply(normalize_postcode)
    df1["n_house"]       = df1["house_number"].apply(normalize_text)
    df1["n_street"]      = df1["street_number"].apply(normalize_text)

    log("🔗 Fetching Smartsheet IOE data...")
    df2 = fetch_smartsheet_df(ss_token, sheet_id)
    df2.rename(columns={
        "Full Address":    "Address",
        "Postal Code":     "PostalCode",
        "Select Measures": "Selected_Measures"
    }, inplace=True)

    df2["Address"] = df2.apply(
        lambda r: r["Address"] if extract_postcode(r["Address"]) else f'{r["Address"]}, {r["PostalCode"]}',
        axis=1
    )
    df2["Postcode"]      = df2["Address"].apply(extract_postcode)
    df2["house_number"]  = df2["Address"].apply(extract_house_number)
    df2["street_number"] = df2.apply(lambda r: extract_street_number(r["Address"], r["house_number"]), axis=1)
    df2["n_postcode"]    = df2["Postcode"].apply(normalize_postcode)
    df2["n_house"]       = df2["house_number"].apply(normalize_text)
    df2["n_street"]      = df2["street_number"].apply(normalize_text)

    matched_row = None
    for _, row in df2[df2["n_postcode"] == df1.loc[0, "n_postcode"]].iterrows():
        if row["n_house"] == df1.loc[0, "n_house"] and row["n_street"] == df1.loc[0, "n_street"]:
            matched_row = row
            break

    if matched_row is None:
        log(f"⚠️  No address match found in Smartsheet — continuing with blank Smartsheet fields")
        matched_address = selected_measures = property_id_value = sap_potential = sap_current = ""
        conservation_flag = ""
        listed_building   = ""
        radon_risk        = ""
        flood_risk        = ""
        subs2030          = ""
        subs2070          = ""
        wdr_risk          = ""
        rainfall_risk     = ""
        wind_speed_risk   = ""
        wildfire_risk     = ""
    else:
        log(f"✅ ADDRESS MATCH found in Smartsheet")
        matched_address   = matched_row.get("Address", "")
        # Robust read: column may be "Selected_Measures" (after rename),
        # "Select Measures" (original), or any column containing "measure".
        if "Selected_Measures" in matched_row.index:
            selected_measures = matched_row["Selected_Measures"]
        elif "Select Measures" in matched_row.index:
            selected_measures = matched_row["Select Measures"]
        else:
            _mcol = next((c for c in matched_row.index if "measure" in str(c).lower()), None)
            selected_measures = matched_row[_mcol] if _mcol else ""
        property_id_value = matched_row.get("Property ID", "")
        sap_potential     = matched_row.get("Proposed Post SAP Score", "")
        sap_current       = matched_row.get("Pre SAP Score (Full SAP)", "")

        log(f"🏷️ Property ID from IOE sheet: {property_id_value}")

    # -- Property Master risk lookup using Property ID --
    if property_id_value and property_master_sheet_id:
        log("🔗 Fetching Property Master Smartsheet...")
        df_master = fetch_property_master_df(ss_token, property_master_sheet_id)
        master_row = find_property_master_row(df_master, property_id_value)

        if master_row is None:
            log(f"⚠️ No Property Master match found for Property ID: {property_id_value}")
            conservation_flag = ""
            listed_building   = ""
            radon_risk        = ""
            flood_risk        = ""
            subs2030          = ""
            subs2070          = ""
            wdr_risk          = ""
            rainfall_risk     = ""
            wind_speed_risk   = ""
            wildfire_risk     = ""
        else:
            log(f"✅ Property Master match found for Property ID: {property_id_value}")
            conservation_flag = get_safe_value(master_row, CONSERVATION_AREA_COL)
            listed_building   = get_safe_value(master_row, LISTED_BUILDING_COL)
            radon_risk        = get_safe_value(master_row, RADON_RISK_COL)

            flood_risk_raw = get_safe_value(master_row, FLOOD_RISK_COL)
            flood_risk = "No" if flood_risk_raw == "" else flood_risk_raw

            subs2030        = get_safe_value(master_row, SUBS_2030_COL)
            subs2070        = get_safe_value(master_row, SUBS_2070_COL)
            wdr_risk        = get_safe_value(master_row, WDR_RISK_COL)
            rainfall_risk   = get_safe_value(master_row, RAINFALL_RISK_COL)
            wind_speed_risk = get_safe_value(master_row, WIND_SPEED_RISK_COL)
            wildfire_risk   = get_safe_value(master_row, WILDFIRE_RISK_COL)

            log(
                f"🌍 Property Master risks → "
                f"Conservation: {conservation_flag} | "
                f"Listed: {listed_building} | "
                f"Radon: {radon_risk} | "
                f"Flood: {flood_risk}"
            )
    else:
        log("⚠️ Property Master lookup skipped — missing Property ID or Property Master Sheet ID")
        conservation_flag = ""
        listed_building   = ""
        radon_risk        = ""
        flood_risk        = ""
        subs2030          = ""
        subs2070          = ""
        wdr_risk          = ""
        rainfall_risk     = ""
        wind_speed_risk   = ""
        wildfire_risk     = ""

    # -- Coordinates --
    latitude, longitude = extract_lat_lon_from_cr(cr_path)
    if latitude == "" or longitude == "":
        raise ValueError("Latitude/Longitude not found in CR PDF.")
    log(f"🌍 Coordinates: {latitude}, {longitude}")

    # -- EPR additional ratings --
    epr_lines = epr_text.splitlines()
    nums, capture = [], False
    for line in epr_lines:
        if "Additional ratings for your home" in line:
            capture = True
            continue
        if capture:
            lc = line.strip()
            if not lc or "office use only" in lc.lower(): break
            m = re.search(r"\d+(\.\d+)?", lc)
            if m: nums.append(m.group(0))
            if len(nums) >= 10: break

    if len(nums) >= 10:
        extra1      = custom_round(float(nums[0]))
        extra2      = custom_round(float(nums[1]))
        extra1_last = custom_round(float(nums[-2]))
        extra2_last = custom_round(float(nums[-1]))
    else:
        extra1 = extra2 = extra1_last = extra2_last = ""

    # -- SN page 1 data --
    sn_page1_text = "\n".join(read_pdf_page_lines(sn_path, 0))
    m_tenure      = re.search(r"Property\s+Tenure:\s*([^\n\r]+)", sn_page1_text)
    raw_tenure    = m_tenure.group(1).strip() if m_tenure else ""
    tenure        = normalise_tenure(raw_tenure)
    date_built    = extract_date_built_from_sn(sn_path)
    num_storeys   = extract_num_storeys_from_sn(sn_path)
    wall_type     = safe_search(r"Type\s+[A-Z]+\s+(.*)", sn_text)
    property_type = extract_property_type_from_epr(epr_text)
    exposure_zone = extract_exposure_zone_from_cr(cr_path)
    tfa           = safe_search(r"Total Floor Area\s*([\d\.]+\s*m²?)", epr_text)
    if tfa:        tfa = tfa.replace("m²", "m2")

    vulnerable_flag                                       = extract_vulnerable_flag_from_cr(cr_path)
    property_constraints_val, property_construction_flag  = extract_constraints_and_construction_from_cr(cr_path)
    orientation_direction                                 = get_building_orientation_direction(cr_path)
    highest_bedroom = extract_highest_bedroom_from_contents(read_pdf_page_lines(cr_path, 1))
    log(f"🏠 Bedrooms: {highest_bedroom} | TFA: {tfa} | Tenure: {tenure}")

    # -- Fetch client spec sheet --
    log("🔗 Fetching Smartsheet Client Specification...")
    df_specs = fetch_client_spec_df(ss_token, second_sheet_id)

    # -- Validate spec columns exist (warn only) --
    if not df_specs.empty:
        if prelim_spec_col and prelim_spec_col not in df_specs.columns:
            log(f"⚠️ Prelim spec column '{prelim_spec_col}' not found in Client Spec sheet — Prelim specs will be blank")
        if construction_spec_col and construction_spec_col not in df_specs.columns:
            log(f"⚠️ Construction spec column '{construction_spec_col}' not found in Client Spec sheet — Construction specs will be blank")

    # -- Bundle all extracted data --
    wb_data = {
        "dwelling_address":           dwelling_address,
        "property_id_value":          property_id_value,
        "local_authority_name":       local_authority_name,   # -> 4.0 B5 only
        "sap_current":                sap_current,
        "sap_potential":              sap_potential,
        "tenure":                     tenure,
        "date_built":                 date_built,
        "wall_type":                  wall_type,
        "property_type":              property_type,
        "num_storeys":                num_storeys,
        "highest_bedroom":            highest_bedroom,
        "exposure_zone":              exposure_zone,
        "property_construction_flag": property_construction_flag,
        "conservation_flag":          conservation_flag,
        "listed_building":            listed_building,
        "property_constraints_val":   property_constraints_val,
        "vulnerable_flag":            vulnerable_flag,
        "tfa_sqm":                    tfa.replace("m", "sqm") if tfa else "",
        "extra1":                     extra1,
        "extra2":                     extra2,
        "extra1_last":                extra1_last,
        "extra2_last":                extra2_last,
        "radon_risk":                 radon_risk,
        "flood_risk":                 flood_risk,
        "subs2030":                   subs2030,
        "subs2070":                   subs2070,
        "wdr_risk":                   wdr_risk,
        "rainfall_risk":              rainfall_risk,
        "wind_speed_risk":            wind_speed_risk,
        "wildfire_risk":              wildfire_risk,
    }

    # ── PRELIM OUTPUT — NOW WITH climate risk, uses PRELIM spec column ──
    log(f"📄 Building Prelim output (spec column: '{prelim_spec_col}')...")
    wb_prelim  = load_workbook(io.BytesIO(template_bytes))
    ws4_prelim = wb_prelim["4.0"]
    write_sheet_4_base(ws4_prelim, wb_data)
    write_sheet_4_climate_risk(ws4_prelim, wb_data)   # NEW: climate risk now included in Prelim too
    ws7_prelim = wb_prelim["7.0"]
    write_sheet_7(ws7_prelim, wb_data, df_specs, selected_measures, orientation_direction, ws4_prelim, spec_col=prelim_spec_col, log=log)
    write_sheet_10(wb_prelim, project_name, sheet10_f3, log=log)
    insert_rooms_with_template(
        cr_pdf_path=cr_path, workbook=wb_prelim, sheet_name="11.0",
        insert_at_row=5, template_row=4, room_col_left="A", room_col_right="J",
        tfa=tfa, highest_bedroom=highest_bedroom, log=log
    )
    prelim_buffer = io.BytesIO()
    wb_prelim.save(prelim_buffer)
    prelim_buffer.seek(0)
    log("✅ Prelim output ready")

    # ── CONSTRUCTION OUTPUT — with climate risk, uses CONSTRUCTION spec column ──
    log(f"🏗️  Building Construction output (spec column: '{construction_spec_col}')...")
    wb_construction  = load_workbook(io.BytesIO(template_bytes))
    ws4_construction = wb_construction["4.0"]
    write_sheet_4_base(ws4_construction, wb_data)
    write_sheet_4_climate_risk(ws4_construction, wb_data)
    ws7_construction = wb_construction["7.0"]
    write_sheet_7(ws7_construction, wb_data, df_specs, selected_measures, orientation_direction, ws4_construction, spec_col=construction_spec_col, log=log)
    write_sheet_10(wb_construction, project_name, sheet10_f3, log=log)
    insert_rooms_with_template(
        cr_pdf_path=cr_path, workbook=wb_construction, sheet_name="11.0",
        insert_at_row=5, template_row=4, room_col_left="A", room_col_right="J",
        tfa=tfa, highest_bedroom=highest_bedroom, log=log
    )
    construction_buffer = io.BytesIO()
    wb_construction.save(construction_buffer)
    construction_buffer.seek(0)
    log("✅ Construction output ready")

    elapsed = time.time() - start_time
    log(f"⏱️  Completed in {elapsed:.1f}s")

    return prelim_buffer, construction_buffer, dwelling_address

# ================================================================
# STREAMLIT UI
# ================================================================

# ── LOGO HEADER ──────────────────────────────────────────────────
if LOGO_B64:
    st.markdown(
        f'''
        <div class="logo-wrap">
            <img src="data:image/png;base64,{LOGO_B64}" alt="Design Specifics" />
            <div class="logo-divider"></div>
            <div>
                <div class="logo-title">Power Asset · Retrofit Design Exporter</div>
                <div class="logo-subtitle">Elmhurst CR · SN · EPR → Excel Design Pack</div>
            </div>
        </div>
        ''',
        unsafe_allow_html=True
    )
else:
    st.markdown(
        '''
        <div class="logo-wrap">
            <div>
                <div class="logo-title">Power Asset · Retrofit Design Exporter</div>
                <div class="logo-subtitle">Logo not found. Place <code>Design_Specifics_logo.png</code> inside the <code>template/</code> folder.</div>
            </div>
        </div>
        ''',
        unsafe_allow_html=True
    )

# ── SIDEBAR — Sheet IDs, Sheet 10 values, Spec columns, Template ──
with st.sidebar:
    st.markdown(
        f'<span class="badge badge-brand">👤 {st.session_state.get("user_email", "")}</span>',
        unsafe_allow_html=True,
    )
    if st.button("Log out"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()
    st.markdown("---")

    st.markdown("### Configuration")
    st.markdown("")

    st.markdown("**Smartsheet IDs**")
    sheet_id = st.text_input("IOE Sheet ID", placeholder="74429503768452", key="sheet_id")
    property_master_sheet_id = st.text_input(
        "Property Master Sheet ID",
        placeholder="Paste Property Master Smartsheet ID",
        key="property_master_sheet_id"
    )
    second_sheet_id = st.text_input("Client Spec Sheet ID", placeholder="1408360446560132", key="second_sheet_id")

    st.markdown("---")

    st.markdown("**Local Authority**")
    local_authority_name = st.text_input(
        "Local Authority Name (4.0 B5)",
        placeholder="e.g. Croydon Council",
        key="local_authority",
        help="Goes to Sheet 4.0 cell B5 only.",
    )

    st.markdown("---")

    st.markdown("**Sheet 10 — Project Details**")
    project_name = st.text_input(
        "Project Name (10.0 B5)",
        placeholder="e.g. Croydon Retrofit Phase 1",
        key="project_name",
        help="Goes to Sheet 10.0 cell B5.",
    )
    sheet10_f3 = st.text_input("REF (10.0 F3)", placeholder="e.g. 1024/CDN", key="f3")

    st.markdown("---")

    st.markdown("**Client Spec — Columns**")
    st.caption("Enter the exact Smartsheet column name for each output. The Prelim report pulls specs from the Prelim column; the Construction report from the Construction column.")

    prelim_spec_col = st.text_input(
        "Prelim column",
        placeholder="e.g. Client Specification",
        key="prelim_spec_col",
        help="Exact Smartsheet column name used for the Prelim report's specifications.",
    )
    construction_spec_col = st.text_input(
        "Construction column",
        placeholder="e.g. Construction Design Specification Notes",
        key="construction_spec_col",
        help="Exact Smartsheet column name used for the Construction report's specifications.",
    )

    prelim_spec_col = prelim_spec_col.strip()
    construction_spec_col = construction_spec_col.strip()

    st.markdown("---")

    st.markdown("**Excel Template**")
    if BUILTIN_TEMPLATE_PATH.exists():
        st.markdown('<span class="badge">✓ Built-in template loaded</span>', unsafe_allow_html=True)
        st.caption(f"Using: {BUILTIN_TEMPLATE_PATH.name}")
    else:
        st.markdown('<span class="badge badge-warn">⚠ Template missing</span>', unsafe_allow_html=True)
        st.caption("Place DESIGN EXPORT TEMPLATE.xlsx inside the assets folder.")

# ── MAIN AREA ────────────────────────────────────────────────────
col_left, col_right = st.columns([1.15, 0.85], gap="large")

with col_left:
    # ── INPUTS CARD: 3 PDFs together ──
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-label">Inputs</div>', unsafe_allow_html=True)
    st.markdown('<div class="card-hint">Upload the three Elmhurst PDFs. The Smartsheet token and Excel template are loaded automatically from the app setup.</div>', unsafe_allow_html=True)

    # Three PDF uploaders side by side
    c1, c2, c3 = st.columns(3)
    with c1:
        cr_file  = st.file_uploader("CR PDF", type=["pdf"], help="Condition Report")
    with c2:
        sn_file  = st.file_uploader("SN PDF", type=["pdf"], help="Survey Notes")
    with c3:
        epr_file = st.file_uploader("EPR PDF", type=["pdf"], help="Energy Performance Report")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── CHECKLIST ──
    ready_flags = {
        "Smartsheet Token":      bool(SMARTSHEET_API_TOKEN),
        "Built-in Excel Template": BUILTIN_TEMPLATE_PATH.exists(),
        "CR PDF":                cr_file is not None,
        "SN PDF":                sn_file is not None,
        "EPR PDF":               epr_file is not None,
        "IOE Sheet ID":          bool(sheet_id),
        "Property Master Sheet ID": bool(property_master_sheet_id),
        "Client Spec Sheet ID":  bool(second_sheet_id),
        "Local Authority":       bool(local_authority_name),
        "Project Name":          bool(project_name),
        "Prelim column":         bool(prelim_spec_col),
        "Construction column":   bool(construction_spec_col),
    }

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-label">Pre-flight Checklist</div>', unsafe_allow_html=True)

    chk_cols = st.columns(2)
    for i, (label, ok) in enumerate(ready_flags.items()):
        icon = "✓" if ok else "○"
        color = "#5B8F73" if ok else "#A0A6B3"
        cls = "ok" if ok else ""
        chk_cols[i % 2].markdown(
            f'<div class="check-item {cls}">'
            f'<span style="color:{color};font-weight:700;font-size:1rem">{icon}</span>'
            f'<span>{label}</span>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown('</div>', unsafe_allow_html=True)

    all_ready = all(ready_flags.values())

    run_btn = st.button("Generate Excel Reports", disabled=not all_ready)

with col_right:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-label">Processing Log</div>', unsafe_allow_html=True)
    log_placeholder = st.empty()
    log_placeholder.markdown('<div class="log-box"><span class="log-empty">Waiting to start…</span></div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    result_placeholder = st.empty()

# ── PROCESS ──────────────────────────────────────────────────────
if run_btn:
    log_lines = []

    def log(msg):
        log_lines.append(msg)
        html_lines = []
        for line in log_lines[-60:]:
            if line.startswith("✅") or line.startswith("✓"):
                cls = "log-ok"
            elif line.startswith("⚠️") or line.startswith("ℹ️"):
                cls = "log-warn"
            elif line.startswith("❌"):
                cls = "log-err"
            else:
                cls = "log-info"
            escaped = line.replace("<", "&lt;").replace(">", "&gt;")
            html_lines.append(f'<div class="{cls}">{escaped}</div>')
        log_placeholder.markdown(
            f'<div class="log-box">{"".join(html_lines)}</div>',
            unsafe_allow_html=True
        )

    try:
        # Save uploaded PDFs to temp files
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(cr_file.read())
            cr_path = f.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(sn_file.read())
            sn_path = f.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(epr_file.read())
            epr_path = f.name

        template_bytes = load_builtin_template_bytes()

        log("🚀 Starting processing…")

        prelim_buf, construction_buf, dwelling_address = process_property(
            cr_path=cr_path,
            sn_path=sn_path,
            epr_path=epr_path,
            template_bytes=template_bytes,
            ss_token=SMARTSHEET_API_TOKEN,
            sheet_id=int(sheet_id),
            property_master_sheet_id=int(property_master_sheet_id),
            second_sheet_id=int(second_sheet_id),
            local_authority_name=local_authority_name,
            project_name=project_name,
            sheet10_f3=sheet10_f3,
            prelim_spec_col=prelim_spec_col,
            construction_spec_col=construction_spec_col,
            log=log,
        )

        log("🎉 Both outputs generated successfully!")

        st.session_state["prelim_buf"]       = prelim_buf.getvalue()
        st.session_state["construction_buf"] = construction_buf.getvalue()
        st.session_state["dwelling_address"] = dwelling_address
        st.session_state["results_ready"]    = True

        for p in [cr_path, sn_path, epr_path]:
            try: os.unlink(p)
            except: pass

    except Exception as e:
        import traceback
        log(f"❌ ERROR: {e}")
        log(traceback.format_exc())
        st.session_state["results_ready"] = False

# ── DOWNLOAD SECTION ──────────────────────────────────────────────
if st.session_state.get("results_ready"):
    addr = st.session_state.get("dwelling_address", "property")
    safe = re.sub(r"[^\w\s-]", "", addr).strip().replace(" ", "_") or "property"

    with result_placeholder.container():
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-label">Download Outputs</div>', unsafe_allow_html=True)
        st.markdown(f'<span class="badge badge-brand">📍 {addr}</span>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        dcol1, dcol2 = st.columns(2)

        with dcol1:
            st.markdown("""
            <div class="output-mini">
                <div class="output-mini-icon">📄</div>
                <div class="output-mini-title">Prelim Report</div>
                <div class="output-mini-sub">Includes climate risk data</div>
            </div>
            """, unsafe_allow_html=True)
            st.download_button(
                label="⬇ Download Prelim",
                data=st.session_state["prelim_buf"],
                file_name=f"{safe}_Prelim.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_prelim",
                use_container_width=True,
            )

        with dcol2:
            st.markdown("""
            <div class="output-mini">
                <div class="output-mini-icon">🏗️</div>
                <div class="output-mini-title">Construction Report</div>
                <div class="output-mini-sub">Includes climate risk data</div>
            </div>
            """, unsafe_allow_html=True)
            st.download_button(
                label="⬇ Download Construction",
                data=st.session_state["construction_buf"],
                file_name=f"{safe}_Construction.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_construction",
                use_container_width=True,
            )

        st.markdown('</div>', unsafe_allow_html=True)

# ── FOOTER ────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<div style="text-align:center;color:#A0A6B3;font-size:0.78rem;padding:0.5rem 0">'
    'Design Specifics · Design Matters'
    '</div>',
    unsafe_allow_html=True
)
